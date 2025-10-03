import os
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from sqlalchemy.exc import SQLAlchemyError
from config import get_engine
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter

# Папки
CHARTS_DIR = "charts"
EXPORTS_DIR = "exports"
os.makedirs(CHARTS_DIR, exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)

# ---------- helper: load queries.sql ----------
def load_named_queries(path="queries.sql"):
    queries = {}
    current_name = None
    current_lines = []
    with open(path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.rstrip("\n")
            if line.strip().lower().startswith("-- name:"):
                # сохранить предыдущий блок
                if current_name:
                    queries[current_name] = "\n".join(current_lines).strip()
                current_name = line.split(":", 1)[1].strip()
                current_lines = []
            else:
                current_lines.append(line)
        if current_name:
            queries[current_name] = "\n".join(current_lines).strip()
    return queries

# ---------- plotting helpers ----------
plt.style.use('seaborn-v0_8')  # аккуратный стиль (не принудителен)
def save_fig(fig, filename):
    path = os.path.join(CHARTS_DIR, filename)
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)
    print(f"Saved chart -> {path}")

def plot_pie(df, idx_col, val_col, title, fname):
    fig, ax = plt.subplots(figsize=(7,7))
    series = df.set_index(idx_col)[val_col]
    series.plot.pie(autopct="%1.1f%%", ax=ax)
    ax.set_ylabel("")
    ax.set_title(title)
    save_fig(fig, fname)

def plot_bar(df, x, y, title, fname, rotate_xticks=45):
    fig, ax = plt.subplots(figsize=(10,6))
    df.plot.bar(x=x, y=y, legend=False, ax=ax)
    ax.set_title(title)
    ax.set_xlabel("")
    ax.set_ylabel(y)
    plt.xticks(rotation=rotate_xticks, ha='right')
    save_fig(fig, fname)

def plot_barh(df, x, y, title, fname):
    fig, ax = plt.subplots(figsize=(9,6))
    df.plot.barh(x=x, y=y, legend=False, ax=ax)
    ax.set_title(title)
    ax.set_xlabel(y)
    save_fig(fig, fname)

def plot_line(df, x, y, title, fname):
    fig, ax = plt.subplots(figsize=(10,5))
    df.plot.line(x=x, y=y, marker="o", ax=ax)
    ax.set_title(title)
    ax.set_xlabel(x)
    ax.set_ylabel(y)
    save_fig(fig, fname)

def plot_hist(df, col, title, fname, bins=20):
    fig, ax = plt.subplots(figsize=(8,5))
    df[col].plot.hist(bins=bins, rwidth=0.9, ax=ax)
    ax.set_title(title)
    ax.set_xlabel(col)
    save_fig(fig, fname)

def plot_scatter(df, x, y, title, fname):
    fig, ax = plt.subplots(figsize=(8,6))
    df.plot.scatter(x=x, y=y, ax=ax, alpha=0.6)
    ax.set_title(title)
    save_fig(fig, fname)

# ---------- export to Excel with formatting ----------
def export_to_excel(dataframes_dict, filename):
    path = os.path.join(EXPORTS_DIR, filename)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.book  # flush

    # Открываем для форматирования
    wb = load_workbook(path)
    total_rows = 0
    for sheet_name, df in dataframes_dict.items():
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column
        total_rows += max(0, len(df))

        # закрепить заголовок
        ws.freeze_panes = "A2"

        # автофильтр
        ws.auto_filter.ref = ws.dimensions

        # применим градиентную заливку для числовых колонок
        numeric_cols = [i+1 for i, dtype in enumerate(df.dtypes) if pd.api.types.is_numeric_dtype(dtype)]
        for col_idx in numeric_cols:
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{max_row}"
            rule = ColorScaleRule(start_type="min", start_color="FFEEEEFF",
                                  mid_type="percentile", mid_value=50, mid_color="FFFFFFAA",
                                  end_type="max", end_color="FFDDFFDD")
            ws.conditional_formatting.add(cell_range, rule)

            # выделение max и min (формулы)
            fill_max = PatternFill(start_color="FF99FF99", end_color="FF99FF99", fill_type="solid")
            fill_min = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
            max_formula = f"={col_letter}2=MAX(${col_letter}$2:${col_letter}${max_row})"
            min_formula = f"={col_letter}2=MIN(${col_letter}$2:${col_letter}${max_row})"
            ws.conditional_formatting.add(cell_range, FormulaRule(formula=[max_formula], fill=fill_max))
            ws.conditional_formatting.add(cell_range, FormulaRule(formula=[min_formula], fill=fill_min))

    wb.save(path)
    print(f"Создан файл {filename}, листов: {len(dataframes_dict)}, строк (всего): {total_rows}")
    return path

# ---------- main ----------
def main():
    engine = get_engine()
    queries = load_named_queries("queries.sql")

    # Список и порядок графиков (ключ в queries.sql -> функция построения)
    try:
        # 1) Pie: жанры после 2000 (использует >=2 JOIN)
        df_pie = pd.read_sql(queries["pie_genres_since_2000"], engine)
        print(f"pie_genres_since_2000: {len(df_pie)} rows")
        plot_pie(df_pie, idx_col="genre", val_col="movie_count",
                 title="Распределение фильмов по жанрам (с 2000 года)", fname="pie_genres_since_2000.png")

        # 2) Bar: топ фильмов + genome tags
        df_bar = pd.read_sql(queries["bar_top_movies_with_genome_tags"], engine)
        print(f"bar_top_movies_with_genome_tags: {len(df_bar)} rows")
        # укоротим подписи
        df_bar['short_title'] = df_bar['title'].str.slice(0,60)
        plot_bar(df_bar, x="short_title", y="total_ratings",
                 title="Топ-10 фильмов по количеству оценок (и их avg_rating)", fname="bar_top_movies.png")

        # 3) Barh: средний рейтинг по жанрам
        df_barh = pd.read_sql(queries["barh_genres_avg_rating_with_tagstats"], engine)
        print(f"barh_genres_avg_rating_with_tagstats: {len(df_barh)} rows")
        plot_barh(df_barh, x="genre", y="avg_rating",
                  title="Средний рейтинг по жанрам (только жанры с >50 фильмов)", fname="barh_genres_avg_rating.png")

        # 4) Line: динамика оценок по годам
        df_line = pd.read_sql(queries["line_ratings_per_year_with_links"], engine)
        print(f"line_ratings_per_year_with_links: {len(df_line)} rows")
        plot_line(df_line, x="year", y="ratings_count",
                  title="Динамика количества оценок по годам (фильмы с внешними ссылками)", fname="line_ratings_years.png")

        # 5) Histogram: распределение средних рейтингов фильмов
        df_hist = pd.read_sql(queries["hist_avg_movie_ratings_with_genome"], engine)
        print(f"hist_avg_movie_ratings_with_genome: {len(df_hist)} rows")
        plot_hist(df_hist, col="avg_rating", title="Распределение средних рейтингов фильмов", fname="hist_avg_movie_ratings.png", bins=25)

        # 6) Scatter: популярность vs качество (и кол-во пользовательских тегов)
        df_scatter = pd.read_sql(queries["scatter_popularity_quality_with_tagcount"], engine)
        print(f"scatter_popularity_quality_with_tagcount: {len(df_scatter)} rows")
        plot_scatter(df_scatter, x="total_ratings", y="avg_rating",
                     title="Популярность vs Качество (только фильмы с >200 оценок)", fname="scatter_popularity_vs_quality.png")

        # Plotly interactive (сохраняем html)
        df_plotly = pd.read_sql(queries["plotly_genre_year_trend"], engine)
        print(f"plotly_genre_year_trend: {len(df_plotly)} rows")
        if not df_plotly.empty:
            fig = px.scatter(df_plotly, x="ratings_count", y="avg_rating",
                             animation_frame="year", animation_group="genre",
                             size="ratings_count", color="genre",
                             hover_name="genre",
                             title="Динамика интереса к жанрам по годам (1995-2015)")
            out_html = os.path.join(CHARTS_DIR, "plotly_genre_year_trend.html")
            fig.write_html(out_html, include_plotlyjs="cdn")
            print(f"Saved interactive Plotly -> {out_html}")
            # fig.show()  # локально можно показать окно

        # Экспорт в Excel: создаём 3 листа (пример)
        exports = {
            "top_movies": df_bar,
            "genres_stats": df_barh,
            "ratings_trend": df_line
        }
        export_filename = "movies_report.xlsx"
        export_path = export_to_excel(exports, export_filename)

    except SQLAlchemyError as e:
        print("Ошибка при выполнении SQL / подключении к БД:", e)
    except KeyError as e:
        print("В queries.sql нет блока с именем:", e)
    except Exception as e:
        print("Неожиданная ошибка:", e)


if __name__ == "__main__":
    main()
